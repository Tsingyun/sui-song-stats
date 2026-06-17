#!/usr/bin/env python3
"""Check if column C has data in the new Excel file."""

import openpyxl

wb = openpyxl.load_workbook('C:/Users/Tsing/Downloads/岁己数据统计 - 副本.xlsx')
ws = wb[wb.sheetnames[0]]

print("检查C列（第3列）的数据：")
print("=" * 60)

audience_count = 0
empty_count = 0
dates_found = []

for i, row in enumerate(ws.iter_rows(min_row=2, max_row=50, values_only=True), 2):
    date_val = row[0]   # A列
    song = row[1]     # B列
    audience = row[2]   # C列
    
    if date_val is not None:
        dates_found.append((i, date_val))
    
    if audience is not None:
        audience_count += 1
        if audience_count <= 10:
            print(f"第{i}行: 日期={date_val}, 歌曲={song}, 观众={audience}")
    elif audience_count == 11:
        print("...")
    
    if audience is None:
        empty_count += 1

print()
print(f"前50行统计：")
print(f"  - C列有数据的行数: {audience_count}")
print(f"  - C列为空的行数: {empty_count}")

print()
print("日期列的前10个有效值：")
for i, (row, val) in enumerate(dates_found[:10], 1):
    print(f"  {i}. 第{row}行: {val}")

# Check a larger sample
print()
print("检查更大的范围（第2-100行）C列有数据的行：")
count = 0
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=100, values_only=True), 2):
    if row[2] is not None:
        count += 1
        if count <= 5:
            print(f"  {i}. 观众={row[2]}, 歌曲={row[1]}")
print(f"  共{count}行有观众数据")
