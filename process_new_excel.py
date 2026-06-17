#!/usr/bin/env python3
"""Process new Excel file (副本.xlsx) - only count column C audience."""

import json, re
from datetime import datetime, timedelta
from collections import Counter

import openpyxl

print("正在读取新Excel文件（副本.xlsx）...")
wb = openpyxl.load_workbook('C:/Users/Tsing/Downloads/岁己数据统计 - 副本.xlsx', data_only=True)
ws = wb[wb.sheetnames[0]]

# Get merged cell ranges for column A (date)
merged_a = {}
for merged_range in ws.merged_cells.ranges:
    if merged_range.min_col == 1 and merged_range.max_col == 1:  # Column A only
        date_val = ws.cell(merged_range.min_row, 1).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            merged_a[row] = date_val

print(f"合并单元格映射：{len(merged_a)} 行")

# Process rows
raw_records = []
current_date = None

for row_idx in range(2, ws.max_row + 1):
    # Get date from merged cell mapping or current row
    if row_idx in merged_a:
        date_val = merged_a[row_idx]
        if date_val is not None:
            if isinstance(date_val, (int, float)):
                base_date = datetime(1899, 12, 30)
                current_date = (base_date + timedelta(days=int(date_val))).strftime('%Y-%m-%d')
            else:
                current_date = str(date_val)[:10] if date_val else None
    # If not in merged_a, use current_date from previous row
    
    song = ws.cell(row_idx, 2).value  # Column B
    audience = ws.cell(row_idx, 3).value  # Column C - cleaned data
    
    if audience is not None and str(audience).strip():
        audience = str(audience).strip()
        if current_date and song:
            raw_records.append({
                'date': current_date,
                'song': str(song).strip(),
                'audience': audience
            })

print(f"处理后记录数: {len(raw_records)}")

# Check Vaserkia count
vaserkia_count = sum(1 for r in raw_records if r['audience'] == 'Vaserkia')
print(f"Vaserkia的点歌次数: {vaserkia_count}")

# Build statistics
# 1. Total leaderboard
audience_counter = Counter(r['audience'] for r in raw_records)
total_leaderboard = [
    {'rank': i+1, 'audience': aud, 'count': cnt}
    for i, (aud, cnt) in enumerate(audience_counter.most_common())
]

# 2. Monthly leaderboard
monthly_data = {}
for r in raw_records:
    month = r['date'][:7]  # YYYY-MM
    if month not in monthly_data:
        monthly_data[month] = Counter()
    monthly_data[month][r['audience']] += 1

monthly_leaderboard = {}
for month in monthly_data:
    monthly_leaderboard[month] = [
        {'rank': i+1, 'audience': aud, 'count': cnt}
        for i, (aud, cnt) in enumerate(monthly_data[month].most_common())
    ]

# 3. Quarterly leaderboard
quarterly_data = {}
for r in raw_records:
    d = r['date']
    y, m = int(d[:4]), int(d[5:7])
    q = f"{y}-Q{(m-1)//3 + 1}"
    if q not in quarterly_data:
        quarterly_data[q] = Counter()
    quarterly_data[q][r['audience']] += 1

quarterly_leaderboard = {}
for q in quarterly_data:
    quarterly_leaderboard[q] = [
        {'rank': i+1, 'audience': aud, 'count': cnt}
        for i, (aud, cnt) in enumerate(quarterly_data[q].most_common())
    ]

# 4. Yearly leaderboard
yearly_data = {}
for r in raw_records:
    y = r['date'][:4]
    if y not in yearly_data:
        yearly_data[y] = Counter()
    yearly_data[y][r['audience']] += 1

yearly_leaderboard = {}
for y in yearly_data:
    yearly_leaderboard[y] = [
        {'rank': i+1, 'audience': aud, 'count': cnt}
        for i, (aud, cnt) in enumerate(yearly_data[y].most_common())
    ]

# 5. Song leaderboard
song_counter = Counter(r['song'] for r in raw_records)
song_leaderboard = [
    {'rank': i+1, 'song': song, 'count': cnt}
    for i, (song, cnt) in enumerate(song_counter.most_common())
]

# 6. Audience details
audience_details = {}
for r in raw_records:
    aud = r['audience']
    if aud not in audience_details:
        audience_details[aud] = {}
    song = r['song']
    if song not in audience_details[aud]:
        audience_details[aud] = {'count': 0, 'dates': []}
    audience_details[aud][song] += 1
    audience_details[aud]['dates'].append(r['date'])

# 7. Trends
trends = {}
for r in raw_records:
    month = r['date'][:7]
    trends[month] = trends.get(month, 0) + 1

# Build output
output = {
    'metadata': {
        'total_records': len(raw_records),
        'date_range': f"{raw_records[0]['date']} 至 {raw_records[-1]['date']}" if raw_records else "",
        'unique_audiences': len(audience_counter),
        'unique_songs': len(song_counter)
    },
    'total_leaderboard': total_leaderboard,
    'monthly_leaderboard': monthly_leaderboard,
    'quarterly_leaderboard': quarterly_leaderboard,
    'yearly_leaderboard': yearly_leaderboard,
    'song_leaderboard': song_leaderboard,
    'audience_details': audience_details,
    'trends': dict(sorted(trends.items())),
    'raw_data': raw_records
}

# Save JSON
with open('song_data_processed.json', 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print(f"\n✅ 数据处理完成（只统计C列）！")
print(f"  - 总点歌记录: {len(raw_records)}")
print(f"  - 独立观众数: {len(audience_counter)}")
print(f"  - 独立歌曲数: {len(song_counter)}")
print(f"  - Vaserkia的点歌次数: {vaserkia_count}")
print(f"\n数据已保存到 song_data_processed.json")
