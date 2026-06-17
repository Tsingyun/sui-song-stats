#!/usr/bin/env python3
"""Read the new Excel file to understand data structure."""

import openpyxl

wb = openpyxl.load_workbook('C:/Users/Tsing/Downloads/岁己数据统计 - 副本.xlsx')
print('工作表列表:', wb.sheetnames)

# Read first (and only) sheet
ws = wb[wb.sheetnames[0]]
print(f'最大行数: {ws.max_row}')
print(f'最大列数: {ws.max_column}')
print()

# Print first 15 rows
print('前15行数据:')
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
    print(f'第{i}行: {row[:6]}')

# Check merged cells
print()
print(f'合并单元格数量: {len(list(ws.merged_cells.ranges))}')
print('前15个合并区域:')
for i, m in enumerate(list(ws.merged_cells.ranges)[:15], 1):
    print(f'  {i}. {m}')
