#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
使用openpyxl正确读取Excel文件，处理合并单元格和异常日期格式
"""

import openpyxl
import json
from collections import defaultdict
from datetime import datetime, timedelta

# 读取Excel文件
file_path = r"C:/Users/Tsing/Downloads/岁己数据统计.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)
ws = wb['岁己今天唱什么']

print("使用openpyxl读取Excel文件...")
print(f"工作表名称: {ws.title}")
print(f"数据范围: {ws.dimensions}")

# 处理合并单元格：获取日期列的合并信息
merged_ranges = list(ws.merged_cells.ranges) if ws.merged_cells else []
date_merges = []
for merge_range in merged_ranges:
    # 检查是否包含A列（日期列）
    if merge_range.min_col == 1:  # A列
        date_merges.append((merge_range.min_row, merge_range.max_row))

print(f"\n找到 {len(date_merges)} 个日期合并区域")

# 前向填充日期
current_date = None
processed_data = []
error_count = 0

print("\n处理数据...")
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):  # 从第2行开始（跳过表头）
    # 检查A列（日期）
    date_value = row[0]
    
    if date_value is not None:
        # 有新日期
        current_date = None  # 重置
        
        if isinstance(date_value, (int, float)):
            # Excel日期序列号
            try:
                base_date = datetime(1899, 12, 30)
                current_date = (base_date + timedelta(days=int(date_value))).strftime('%Y-%m-%d')
            except:
                error_count += 1
                continue
        elif isinstance(date_value, str):
            # 字符串格式 - 跳过非日期字符串
            date_str = str(date_value).strip()
            if any(c in date_str for c in ['月', '年', '日', '注', '备', '总']):
                error_count += 1
                continue
            
            # 尝试解析日期
            try:
                for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d']:
                    try:
                        current_date = datetime.strptime(date_str, fmt).strftime('%Y-%m-%d')
                        break
                    except:
                        continue
            except:
                error_count += 1
                continue
        elif hasattr(date_value, 'strftime'):
            # datetime对象
            try:
                current_date = date_value.strftime('%Y-%m-%d')
            except:
                error_count += 1
                continue
        else:
            error_count += 1
            continue
    
    if current_date is None:
        continue
    
    # 获取B列（歌曲名）
    song_name = row[1]
    if song_name is None:
        continue
    
    song_name = str(song_name).strip()
    if not song_name:
        continue
    
    # 收集所有点歌的观众
    audiences = []
    
    # C列：哪个饼子点到的？
    primary_audience = row[2]
    if primary_audience is not None:
        audience_name = str(primary_audience).strip()
        if audience_name:
            audiences.append(audience_name)
    
    # E列及之后：参与拼好歌的饼子
    for col_idx in range(4, len(row)):  # 从E列（索引4）开始
        audience = row[col_idx]
        if audience is not None:
            audience_name = str(audience).strip()
            if audience_name:
                audiences.append(audience_name)
    
    # 去重
    audiences = list(set(audiences))
    
    # 为每位观众创建一条记录
    for audience in audiences:
        if audience:
            processed_data.append({
                'date': current_date,
                'song': song_name,
                'audience': audience,
                'month': current_date[:7],
                'quarter': f"{current_date[:4]}-Q{((int(current_date[5:7]) - 1) // 3 + 1)}",
                'year': current_date[:4]
            })

print(f"处理后的点歌记录数: {len(processed_data)}")
if error_count > 0:
    print(f"跳过的异常行数: {error_count}")

# 验证：统计Vaserkia的点歌次数
vaserkia_count = sum(1 for r in processed_data if r['audience'] == 'Vaserkia')
print(f"\n验证 - Vaserkia的点歌次数: {vaserkia_count}")

# 统计
audience_total = defaultdict(int)
audience_songs = defaultdict(set)
song_count = defaultdict(int)

for req in processed_data:
    audience_total[req['audience']] += 1
    audience_songs[req['audience']].add(req['song'])
    song_count[req['song']] += 1

# 按月统计
monthly_stats = defaultdict(lambda: defaultdict(int))
for req in processed_data:
    monthly_stats[req['month']][req['audience']] += 1

# 按季度统计
quarterly_stats = defaultdict(lambda: defaultdict(int))
for req in processed_data:
    quarterly_stats[req['quarter']][req['audience']] += 1

# 按年度统计
yearly_stats = defaultdict(lambda: defaultdict(int))
for req in processed_data:
    yearly_stats[req['year']][req['audience']] += 1

# 观众喜好
audience_preferences = defaultdict(lambda: defaultdict(int))
for req in processed_data:
    audience_preferences[req['audience']][req['song']] += 1

# 准备输出数据
output_data = {
    'metadata': {
        'total_records': len(processed_data),
        'date_range': f"{min(r['date'] for r in processed_data)} 至 {max(r['date'] for r in processed_data)}",
        'unique_audiences': len(audience_total),
        'unique_songs': len(song_count)
    },
    'total_leaderboard': [
        {'rank': i+1, 'audience': a, 'count': c, 'unique_songs': len(audience_songs[a])}
        for i, (a, c) in enumerate(sorted(audience_total.items(), key=lambda x: -x[1])[:50])
    ],
    'song_leaderboard': [
        {'rank': i+1, 'song': s, 'count': c}
        for i, (s, c) in enumerate(sorted(song_count.items(), key=lambda x: -x[1])[:50])
    ],
    'monthly_leaderboard': {},
    'quarterly_leaderboard': {},
    'yearly_leaderboard': {},
    'audience_preferences': {},
    'monthly_trends': defaultdict(int),
    'raw_data': processed_data
}

# 月度排行榜
for month in sorted(monthly_stats.keys()):
    top_audiences = sorted(monthly_stats[month].items(), key=lambda x: -x[1])[:10]
    output_data['monthly_leaderboard'][month] = [
        {'rank': i+1, 'audience': a, 'count': c}
        for i, (a, c) in enumerate(top_audiences)
    ]

# 季度排行榜
for quarter in sorted(quarterly_stats.keys()):
    top_audiences = sorted(quarterly_stats[quarter].items(), key=lambda x: -x[1])[:10]
    output_data['quarterly_leaderboard'][quarter] = [
        {'rank': i+1, 'audience': a, 'count': c}
        for i, (a, c) in enumerate(top_audiences)
    ]

# 年度排行榜
for year in sorted(yearly_stats.keys()):
    top_audiences = sorted(yearly_stats[year].items(), key=lambda x: -x[1])[:10]
    output_data['yearly_leaderboard'][year] = [
        {'rank': i+1, 'audience': a, 'count': c}
        for i, (a, c) in enumerate(top_audiences)
    ]

# 观众喜好
for audience in list(audience_total.keys())[:20]:
    top_songs = sorted(audience_preferences[audience].items(), key=lambda x: -x[1])[:5]
    output_data['audience_preferences'][audience] = [
        {'song': s, 'count': c}
        for s, c in top_songs
    ]

# 月度趋势
for req in processed_data:
    output_data['monthly_trends'][req['month']] += 1

# 保存数据
with open('song_data_processed.json', 'w', encoding='utf-8') as f:
    json.dump(output_data, f, ensure_ascii=False, indent=2)

print("\n✅ 数据处理完成（使用openpyxl）！")
print(f"  - 总点歌记录: {output_data['metadata']['total_records']}")
print(f"  - 独立观众数: {output_data['metadata']['unique_audiences']}")
print(f"  - 独立歌曲数: {output_data['metadata']['unique_songs']}")
print(f"  - Vaserkia的点歌次数: {audience_total.get('Vaserkia', 0)}")
print("\n数据已保存到 song_data_processed.json")
